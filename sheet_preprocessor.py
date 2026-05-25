"""Sheet 预处理器 — 去空行空列、类型检测、层级结构提取。

将 openpyxl worksheet 转换为干净的中间文本，供 AI 重排版使用。
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from enum import Enum

import openpyxl


class SheetType(Enum):
    DATA_TABLE = "data_table"
    CONFIG_TABLE = "config_table"
    ENUM_TABLE = "enum_table"
    VERSION_LOG = "version_log"
    DOCUMENT = "document"


@dataclass
class ImageInfo:
    index: int          # 图片在 sheet 中的序号
    from_row: int       # 锚点行（0-based）
    from_col: int       # 锚点列
    data: bytes         # 图片二进制数据
    fmt: str            # "png" / "jpeg" / "gif"


@dataclass
class CleanedSheet:
    name: str
    sheet_type: SheetType
    content: str
    row_count: int
    estimated_tokens: int
    images: list[ImageInfo] = field(default_factory=list)


# ---------- 合并单元格处理 ----------

def _unmerge_and_fill(ws) -> None:
    """取消合并单元格并用左上角值填充。"""
    for merge_range in list(ws.merged_cells.ranges):
        min_row, min_col = merge_range.min_row, merge_range.min_col
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


# ---------- 图片提取 ----------

_MAGIC_BYTES = [
    (b'\x89PNG', 'png'),
    (b'\xff\xd8\xff', 'jpeg'),
    (b'GIF8', 'gif'),
    (b'BM', 'bmp'),
]


def _detect_image_format(data: bytes) -> str:
    """通过 magic bytes 检测图片格式。"""
    for magic, fmt in _MAGIC_BYTES:
        if data[:len(magic)] == magic:
            return fmt
    if data[:4] == b'RIFF' and data[8:12] == b'WEBP':
        return 'webp'
    return 'png'  # fallback


def _extract_images(ws) -> list[ImageInfo]:
    """从 worksheet 中提取所有嵌入图片。"""
    images = []
    ws_images = getattr(ws, '_images', None)
    if not ws_images:
        return images

    for idx, img in enumerate(ws_images):
        try:
            data = img._data()
            if not data:
                continue
            # 处理 file-like 对象
            if hasattr(data, 'read'):
                data.seek(0)
                data = data.read()
            elif hasattr(data, 'getvalue'):
                data = data.getvalue()
            if isinstance(data, (bytearray, memoryview)):
                data = bytes(data)

            fmt = _detect_image_format(data)
            anchor = img.anchor
            from_obj = getattr(anchor, '_from', None)
            from_row = getattr(from_obj, 'row', 0) if from_obj else 0
            from_col = getattr(from_obj, 'col', 0) if from_obj else 0
            images.append(ImageInfo(
                index=idx,
                from_row=from_row,
                from_col=from_col,
                data=data,
                fmt=fmt,
            ))
        except Exception:
            continue  # 跳过有问题的图片

    return images


# ---------- 读取与清洗 ----------

def _read_raw(ws) -> list[list[str]]:
    """读取 worksheet 为 list[list[str]]，每个元素是单元格文本。"""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v).strip() if v is not None else "" for v in row])
    return rows


def _find_empty_cols(rows: list[list[str]]) -> set[int]:
    """找出全空的列索引集合。"""
    if not rows:
        return set()
    max_cols = max(len(r) for r in rows)
    empty = set()
    for c in range(max_cols):
        if all(c >= len(r) or not r[c] for r in rows):
            empty.add(c)
    return empty


def _clean_rows(rows: list[list[str]]) -> tuple[list[list[str]], list[int], list[int]]:
    """去空列+去连续空行。返回 (干净行, 保留列的原始索引列表, 各行对应的原始行号)。

    保留列的原始索引用于后续层级提取。
    各行对应的原始行号用于图片占位符定位。
    """
    if not rows:
        return [], [], []

    empty_cols = _find_empty_cols(rows)
    max_cols = max(len(r) for r in rows)
    kept_cols = [c for c in range(max_cols) if c not in empty_cols]

    # 去空列
    stripped = []
    for row in rows:
        new_row = [row[c] if c < len(row) else "" for c in kept_cols]
        stripped.append(new_row)

    # 去连续空行（最多保留1个），去首尾空行
    result = []
    result_indices = []
    prev_empty = False
    for idx, row in enumerate(stripped):
        is_empty = not any(row)
        if is_empty:
            if not prev_empty:
                result.append(row)
                result_indices.append(idx)
            prev_empty = True
        else:
            prev_empty = False
            result.append(row)
            result_indices.append(idx)

    while result and not any(result[0]):
        result.pop(0)
        result_indices.pop(0)
    while result and not any(result[-1]):
        result.pop()
        result_indices.pop()

    return result, kept_cols, result_indices


# ---------- 类型检测 ----------

def _is_numeric(s: str) -> bool:
    s = s.strip().replace(",", "").replace("%", "").replace("¥", "").replace("$", "")
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


_VERSION_KEYWORDS = ("版本", "更新", "变更", "changelog", "version", "log", "记录")
_CONFIG_KEYWORDS = ("配置", "config", "参数", "parameter", "setting", "设置")
_ENUM_KEYWORDS = ("枚举", "enum", "常量", "const", "类型", "type", "定义")


def _detect_type(sheet_name: str, rows: list[list[str]]) -> SheetType:
    name_lower = sheet_name.lower()

    for kw in _VERSION_KEYWORDS:
        if kw in name_lower:
            return SheetType.VERSION_LOG
    for kw in _CONFIG_KEYWORDS:
        if kw in name_lower:
            return SheetType.CONFIG_TABLE
    for kw in _ENUM_KEYWORDS:
        if kw in name_lower:
            return SheetType.ENUM_TABLE

    if not rows:
        return SheetType.DOCUMENT

    cell_counts = []
    numeric_rows = 0
    total_rows = 0

    for row in rows:
        non_empty = [c for c in row if c]
        if not non_empty:
            continue
        total_rows += 1
        cell_counts.append(len(non_empty))
        num_count = sum(1 for c in non_empty if _is_numeric(c))
        if num_count / len(non_empty) > 0.5:
            numeric_rows += 1

    if not total_rows:
        return SheetType.DOCUMENT

    max_cols = max(len(r) for r in rows) if rows else 0

    # 配置表检测：列数少（≤3），第一列大多是文本标签
    if max_cols <= 3 and total_rows >= 3:
        first_col_texts = sum(
            1 for r in rows if r and r[0] and not _is_numeric(r[0])
        )
        if first_col_texts / total_rows > 0.7:
            return SheetType.CONFIG_TABLE

    # 枚举表检测：有 ID 列（首列多为数字）+ 文本表头
    if max_cols >= 2 and total_rows >= 3:
        header = rows[0] if rows else []
        non_empty_header = [c for c in header if c]
        header_all_text = non_empty_header and all(
            not _is_numeric(c) for c in non_empty_header
        )
        if header_all_text:
            data_rows = [r for r in rows[1:] if any(r)]
            if data_rows:
                first_col_numeric = sum(
                    1 for r in data_rows if r and _is_numeric(r[0])
                )
                if first_col_numeric / len(data_rows) > 0.7:
                    return SheetType.ENUM_TABLE

    # 原有数据表检测
    if numeric_rows / total_rows > 0.4:
        return SheetType.DATA_TABLE

    avg_cells = sum(cell_counts) / len(cell_counts)
    if avg_cells <= 3:
        return SheetType.DOCUMENT

    if cell_counts:
        mode_count = max(set(cell_counts), key=cell_counts.count)
        consistency = cell_counts.count(mode_count) / len(cell_counts)
        if consistency > 0.5 and mode_count >= 3:
            return SheetType.DATA_TABLE

    return SheetType.DOCUMENT


# ---------- 图片占位符映射 ----------

def _build_image_map(rows, images, orig_row_indices, sheet_name):
    """构建图片插入映射：cleaned_row_index -> 占位符列表。

    key=-1 表示图片在所有内容行之前。
    """
    img_after = defaultdict(list)
    if not images or not orig_row_indices:
        return img_after

    for img_info in sorted(images, key=lambda x: (x.from_row, x.index)):
        insert_after = -1
        for ci in range(len(rows)):
            if (ci < len(orig_row_indices)
                    and orig_row_indices[ci] <= img_info.from_row
                    and any(rows[ci])):
                insert_after = ci
        img_after[insert_after].append(
            f"{{{{IMG:{sheet_name}:{img_info.index}}}}}"
        )

    return img_after


# ---------- 内容格式化 ----------

def _hierarchical_text(rows: list[list[str]], orig_cols: list[int],
                       images=None, orig_row_indices=None, sheet_name="") -> str:
    """利用原始列索引推断层级结构。

    原始列索引越大 → 层级越深 → 输出 2空格缩进。
    """
    if not rows:
        return ""

    # 收集所有出现过内容的原始列索引
    used_orig_cols: set[int] = set()
    for row in rows:
        for i, cell in enumerate(row):
            if cell and i < len(orig_cols):
                used_orig_cols.add(orig_cols[i])

    if not used_orig_cols:
        return ""

    sorted_used = sorted(used_orig_cols)
    col_to_level = {col: lvl for lvl, col in enumerate(sorted_used)}

    # 图片占位符映射
    img_after = _build_image_map(rows, images, orig_row_indices, sheet_name)

    lines: list[str] = []

    # 所有内容之前的图片
    for ph in img_after.get(-1, []):
        lines.append(ph)

    for row_idx, row in enumerate(rows):
        if not any(row):
            if lines and lines[-1] != "":
                lines.append("")
            continue

        # 找第一个非空单元格确定层级
        first_orig_col = None
        parts: list[str] = []
        for i, cell in enumerate(row):
            if cell and i < len(orig_cols):
                if first_orig_col is None:
                    first_orig_col = orig_cols[i]
                parts.append(cell)

        if not parts:
            continue

        level = col_to_level.get(first_orig_col, 0)
        indent = "  " * level
        combined = " | ".join(parts) if len(parts) > 1 else parts[0]
        lines.append(f"{indent}{combined}")

        # 插入此行之后的图片占位符
        for ph in img_after.get(row_idx, []):
            lines.append(ph)

    return "\n".join(lines)


def _markdown_table(rows: list[list[str]],
                    images=None, orig_row_indices=None, sheet_name="") -> str:
    """格式化为干净的 Markdown 表格。"""
    non_empty_rows = []
    non_empty_cleaned_indices = []
    for ci, r in enumerate(rows):
        if any(r):
            non_empty_rows.append(r)
            non_empty_cleaned_indices.append(ci)

    if not non_empty_rows:
        return ""

    # 图片占位符映射
    img_after = _build_image_map(rows, images, orig_row_indices, sheet_name)

    max_cols = max(len(r) for r in non_empty_rows)
    normalized = [r + [""] * (max_cols - len(r)) for r in non_empty_rows]

    lines = []

    # 所有内容之前的图片
    for ph in img_after.get(-1, []):
        lines.append(ph)

    # 表头
    header = normalized[0]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * max_cols) + " |")
    for ph in img_after.get(non_empty_cleaned_indices[0], []):
        lines.append(ph)

    # 数据行
    for i, row in enumerate(normalized[1:], 1):
        lines.append("| " + " | ".join(row) + " |")
        for ph in img_after.get(non_empty_cleaned_indices[i], []):
            lines.append(ph)

    return "\n".join(lines)


# ---------- 公开接口 ----------

def preprocess_sheet(ws, sheet_name: str) -> CleanedSheet:
    """预处理单个 worksheet。"""
    _unmerge_and_fill(ws)
    raw = _read_raw(ws)
    images = _extract_images(ws)

    if not raw or all(not any(r) for r in raw):
        return CleanedSheet(sheet_name, SheetType.DOCUMENT, "*(空表)*", 0, 5, images)

    cleaned, orig_cols, orig_row_indices = _clean_rows(raw)
    sheet_type = _detect_type(sheet_name, cleaned)

    if sheet_type in (SheetType.DATA_TABLE, SheetType.CONFIG_TABLE, SheetType.ENUM_TABLE):
        content = _markdown_table(cleaned, images, orig_row_indices, sheet_name)
    else:
        content = _hierarchical_text(cleaned, orig_cols, images, orig_row_indices, sheet_name)

    row_count = sum(1 for r in cleaned if any(r))
    estimated_tokens = len(content) // 2

    return CleanedSheet(sheet_name, sheet_type, content, row_count, estimated_tokens, images)


def preprocess_workbook(xlsx_path: str) -> list[CleanedSheet]:
    """预处理整个 xlsx 文件，跳过隐藏 sheet (hidden / veryHidden)。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    results = []
    skipped_hidden = []
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.sheet_state != "visible":
            skipped_hidden.append(f"{name}({ws.sheet_state})")
            continue
        results.append(preprocess_sheet(ws, name))
    if skipped_hidden:
        print(f"  跳过隐藏 sheet: {', '.join(skipped_hidden)}")
    wb.close()
    return results
